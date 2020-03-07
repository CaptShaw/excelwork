#! python3
# -*- coding: utf-8 -*-

__author__ = 'CaptShaw'

"""
    12.13.1 乘法表 - 创建程序 multiplicationTable.py，从命令行接受数字 N，在一个 Excel 电子表格 中创建一个 N×N 的乘法表
"""
import openpyxl, os
from openpyxl.styles import Font

num = 6
os.chdir(r'C:\Users\Shaw\PycharmProjects\excelwork')
wb = openpyxl.workbook.Workbook()
sheet = wb['Sheet']
Font = Font(bold=True)

for i in range(1, num + 1):
    sheet.cell(row=i + 1, column=1).font = Font
    sheet.cell(row=1, column=i + 1).font = Font
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=1, column=i + 1).value = i

for r in range(2, num + 2):
    for c in range(2, num + 2):
        sheet.cell(row=r, column=c).value = sheet.cell(row=1, column=c).value * sheet.cell(row=r, column=1).value
wb.save('result.xlsx')
