#! python3
# -*- coding: utf-8 -*-

__author__ = 'CaptShaw'

"""
    some excel work
    http://openpyxl.readthedocs.org/
"""

import openpyxl

example = r'C:\Users\Shaw\PycharmProjects\excelwork\example\censuspopdata.xlsx'
wb = openpyxl.load_workbook(example)
# print(wb.get_sheet_names())
# in an uptodate way
print(wb.sheetnames)
# sheet = wb.get_sheet_by_name('Sheet1')
# in an uptodate way
sheet = wb.active
print(sheet)
print(sheet.max_column,sheet.max_row)

